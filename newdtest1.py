import os
import glob
import xlrd
import re
from xlutils.copy import copy as xl_copy
from openpyxl import load_workbook

# ───── Constants ─────
DOC_DIR = r"C:\\MDM\\1.mdm-ade-import-definitions\\mdm-ade-import-definitions"
CMDD_FILE = r"C:\\MDM\\Bitbucket\\mdm-cmdd\\CMDD-DataPointDefinition.xls"
DATA_DICT_FILE = r"C:\\MDM\\IngestionScriptingTools\\002_Ingestion Script xls to XML Conversion Tool\\ADE Data Dictionary\\DataDictionary.xlsx"

# ───── Fixed Headers ─────
FIXED_HEADERS = [
    'Ingestion id', 'CMDD Field ID', 
    'DataPoint-0', 'DataPoint-0 Source', 'DataPoint-0 Condition', 'DataPoint-0 DataPointContainerRule',
    'DataPoint-1', 'DataPoint-1 Source', 'DataPoint-1 Condition', 'DataPoint-1 DataPointContainerRule',
    'DataPoint-2', 'DataPoint-2 Source', 'DataPoint-2 Condition', 'DataPoint-2 DataPointContainerRule',
    'DataPoint-3', 'DataPoint-3 Source', 'DataPoint-3 Condition', 'DataPoint-3 DataPointContainerRule',
    'DataPoint-4', 'DataPoint-4 Source', 'DataPoint-4 Condition', 'DataPoint-4 DataPointContainerRule',
    'DataPoint-5', 'DataPoint-5 Source', 'DataPoint-5 Condition', 'DataPoint-5 DataPointContainerRule',
    'DataPoint-6', 'DataPoint-6 Source', 'DataPoint-6 Condition', 'DataPoint-6 DataPointContainerRule',
    'Source Data Point Parent', 'TransformationRule', 'Extraction Rule', 'Default Value', 'Collaborator Association',
    'RELATIONSHIP', 'Document Name', 'relationalDatapoint', 'dataCorrection', 'dataStoreDataPointName', 'formattingRule'
]

# ───── Utilities ─────
def get_latest_doc_file(directory, pattern):
    file_pattern = os.path.join(directory, pattern)
    files = glob.glob(file_pattern)
    if not files:
        raise FileNotFoundError("No matching document files found.")
    return max(files, key=os.path.getmtime)

def read_cmdd_xls_to_dicts(file_path):
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)
    headers = sheet.row_values(0)
    return [dict(zip(headers, sheet.row_values(i))) for i in range(1, sheet.nrows)]

def search_absolute_path_by_keyword(keyword, cmdd_data):
    keyword = keyword.lower()
    return [row for row in cmdd_data if keyword in str(row.get('Absolute Path', '')).lower()]

def select_absolute_path_from_matches(matches):
    for idx, row in enumerate(matches, start=1):
        print(f"{idx}. CMDD Field ID#: {row.get('CMDD Field ID#')} | Absolute Path: {row.get('Absolute Path')}")
    choice = int(input("Select row number: "))
    selected = matches[choice - 1]
    return selected.get('CMDD Field ID#'), selected.get('Absolute Path')

def get_next_ingestion_id(latest_doc_file):
    book = xlrd.open_workbook(latest_doc_file)
    sheet = book.sheet_by_index(0)
    headers = [sheet.cell_value(0, col).strip().lower().replace(" ", "") for col in range(sheet.ncols)]
    idx = next((i for i, h in enumerate(headers) if h in ['ingestionid', 'ingestionfieldid']), None)
    ade_numbers = [
        int(re.match(r'ADE(\d{6})', str(sheet.cell_value(i, idx))).group(1))
        for i in range(1, sheet.nrows)
        if re.match(r'ADE\d{6}', str(sheet.cell_value(i, idx)))
    ]
    next_number = max(ade_numbers) + 1 if ade_numbers else 1
    return f"ADE{next_number:06d}"

def get_collaborator_association(latest_doc_file):
    book = xlrd.open_workbook(latest_doc_file)
    try:
        sheet = book.sheet_by_name("IngestionCollAssociationConfig")
    except:
        return "NA"
    headers = [sheet.cell_value(0, col).strip() for col in range(sheet.ncols)]
    idx = headers.index("Collaborator Association Reference Name")
    values = sorted(set(str(sheet.cell_value(i, idx)).strip() for i in range(1, sheet.nrows) if sheet.cell_value(i, idx)))
    values.append("NA")
    print("Collaborator Association:")
    for i, val in enumerate(values, 1):
        print(f"{i}. {val}")
    choice = int(input("Select Collaborator Association: "))
    return values[choice - 1]

def get_data_point_name_from_xlsx(data_dict_file, field_name_id):
    wb = load_workbook(data_dict_file, data_only=True)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    try:
        field_col_index = headers.index("Field Name (with ID)")
        data_point_col_index = headers.index("Data Point Name")
    except ValueError:
        raise Exception("Required columns not found in DATA_DICT_FILE.")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        field_value = row[field_col_index]
        if field_value and str(field_value).strip().lower() == field_name_id.strip().lower():
            return row[data_point_col_index]

    raise ValueError(f"Field '{field_name_id}' not found in 'Field Name (with ID)' column.")

# ───── Updated Function ─────
def apply_data_point_columns(latest_doc_file, cmdd_id, abs_path, ingestion_id_value,
                             field_name_id, data_point_name, clone_target_dp,
                             collaborator_value, doc_name,
                             transformation_rule_input, default_value_input,
                             sheet_name='IngestionConfig'):

    path_parts = abs_path.strip().split('.')
    print("Select which path parts to tag with instance (optional):")
    for i, part in enumerate(path_parts):
        print(f"{i}: {part}")
    indices = input("Comma-separated indices or Enter to skip: ").strip()
    instance = input("Enter instance number (optional): ").strip()
    if indices and instance:
        idx_list = [int(i.strip()) for i in indices.split(',') if i.strip().isdigit()]
        for i in idx_list:
            path_parts[i] = f"{path_parts[i]}#Instance{instance}"

    book = xlrd.open_workbook(latest_doc_file, formatting_info=True)
    sheet = book.sheet_by_name(sheet_name)
    headers = FIXED_HEADERS
    writable = xl_copy(book)
    ws = writable.get_sheet(book.sheet_names().index(sheet_name))
    row_idx = sheet.nrows
    row_data = ['' for _ in headers]

    def set_val(col, val):
        if col in headers:
            row_data[headers.index(col)] = val

    set_val('Ingestion id', ingestion_id_value)
    set_val('CMDD Field ID', cmdd_id)
    set_val('DataPoint-6 Source', field_name_id)
    set_val(f'{clone_target_dp} Source', data_point_name)
    set_val('Collaborator Association', collaborator_value)
    set_val('Document Name', doc_name)
    set_val('TransformationRule', transformation_rule_input)
    set_val('Default Value', default_value_input)

    if path_parts[0].upper() in ['APPLICANT#INSTANCE2', 'APPLICANT#INSTANCE3', 'APPLICANT#INSTANCE4', 'APPLICANT#INSTANCE5', 'COBORROWER']:
        set_val('DataPoint-0 DataPointContainerRule', 'ApplicantRule')

    for i in range(min(6, len(path_parts) - 1)):
        set_val(f'DataPoint-{i}', path_parts[i])
    set_val('DataPoint-6', path_parts[-1])

    for i in range(len(headers)):
        if row_data[i] is None or str(row_data[i]).strip() == '':
            row_data[i] = 'NA'

    for i, val in enumerate(row_data):
        ws.write(row_idx, i, val)

    writable.save(latest_doc_file)
    print(f"\n✅ Data written to row {row_idx} in {latest_doc_file}")

# ───── Main Execution ─────
# (unchanged main block remains as-is)
if __name__ == '__main__':
    doc_name = input("Enter Document Name (without .xls): ")
    latest_doc_file = get_latest_doc_file(DOC_DIR, doc_name + ".xls")
    print("Found latest file:", latest_doc_file)

    field_name_id = input("Enter the Field Name (with ID): ")

    data_point_name = get_data_point_name_from_xlsx(DATA_DICT_FILE, field_name_id)
    
    keyword = input("Enter keyword to search Absolute Path: ")
    cmdd_data = read_cmdd_xls_to_dicts(CMDD_FILE)
    matches = search_absolute_path_by_keyword(keyword, cmdd_data)
    cmdd_id, abs_path = select_absolute_path_from_matches(matches)

    print("Select which DataPoint-n Source to clone into:")
    for i in range(7):
        print(f"{i}. DataPoint-{i} Source")
    while True:
        try:
            dp_choice = int(input("Enter the number (0–6): "))
            if 0 <= dp_choice <= 6:
                clone_target_dp = f'DataPoint-{dp_choice}'
                break
            else:
                print("Invalid selection. Choose 0 to 6.")
        except ValueError:
            print("Please enter a number.")


    ingestion_id_value = get_next_ingestion_id(latest_doc_file)
    collaborator_value = get_collaborator_association(latest_doc_file)
    transformation_rule_input = input("Enter TransformationRule: ")
    default_value_input = input("Enter Default Value: ")

    apply_data_point_columns(latest_doc_file, cmdd_id, abs_path, ingestion_id_value,
                             field_name_id, data_point_name, clone_target_dp,
                             collaborator_value, doc_name,
                             transformation_rule_input, default_value_input)