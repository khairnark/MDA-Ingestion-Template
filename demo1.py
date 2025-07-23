import os
import glob
import xlrd
from xlutils.copy import copy as xl_copy
import re
from openpyxl import load_workbook
# ─────────── Configurable Paths ───────────
DOC_DIR = r"C:\\MDM\\1.mdm-ade-import-definitions\\mdm-ade-import-definitions"
CMDD_FILE = r"C:\\MDM\\Bitbucket\\mdm-cmdd\\CMDD-DataPointDefinition.xls"
DATA_DICT_FILE = r"C:\\MDM\\IngestionScriptingTools\\002_Ingestion Script xls to XML Conversion Tool\\ADE Data Dictionary\\DataDictionary.xlsx"

# ─────────── Fixed Header ───────────
FIXED_HEADERS = [
    'Ingestion id', 'CMDD Field ID', 
    'DataPoint-0', 'DataPoint-0 Source', 'DataPoint-0 Condition', 'DataPoint-0 DataPointContainerRule',
    'DataPoint-1', 'DataPoint-1 Source', 'DataPoint-1 Condition', 'DataPoint-1 DataPointContainerRule',
    'DataPoint-2', 'DataPoint-2 Source', 'DataPoint-2 Condition', 'DataPoint-2 DataPointContainerRule',
    'DataPoint-3', 'DataPoint-3 Source', 'DataPoint-3 Condition', 'DataPoint-3 DataPointContainerRule',
    'DataPoint-4', 'DataPoint-4 Source', 'DataPoint-4 Condition', 'DataPoint-4 DataPointContainerRule',
    'DataPoint-5', 'DataPoint-5 Source', 'DataPoint-5 Condition', 'DataPoint-5 DataPointContainerRule',
    'DataPoint-6', 'DataPoint-6 Source', 'DataPoint-6 Condition', 'DataPoint-6 DataPointContainerRule',
    'Source Data Point Parent', 'TransformationRule', 'Extraction Rule', 'Default Value', 'Collaborator Association',
    'RELATIONSHIP', 'Document Name', 'relationalDatapoint', 'dataCorrection', 'dataStoreDataPointName', 'formattingRule'
]
# ─────────── Helper Functions ───────────
def get_latest_doc_file(directory, pattern):
    file_pattern = os.path.join(directory, pattern)
    files = glob.glob(file_pattern)
    if not files:
        raise FileNotFoundError("No matching document files found.")
    return max(files, key=os.path.getmtime)

def read_cmdd_xls_to_dicts(file_path):
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)
    headers = sheet.row_values(0)
    return [dict(zip(headers, sheet.row_values(i))) for i in range(1, sheet.nrows)]

def search_absolute_path_by_keyword(keyword, cmdd_data):
    keyword = keyword.lower()
    return [row for row in cmdd_data if keyword in str(row.get('Absolute Path', '')).lower()]

def select_absolute_path_from_matches(matches):
    for idx, row in enumerate(matches, start=1):
        print(f"{idx}. CMDD Field ID#: {row.get('CMDD Field ID#')} | Absolute Path: {row.get('Absolute Path')}")
    choice = int(input("Select row number: "))
    selected = matches[choice - 1]
    return selected.get('CMDD Field ID#'), selected.get('Absolute Path')

def get_source_data_point(field_name, data_dict_path):
    ext = os.path.splitext(data_dict_path)[1].lower()

    if ext == ".xls":
        # Use xlrd for .xls
        book = xlrd.open_workbook(data_dict_path)
        sheet = book.sheet_by_name("Data_Dictionary")
        headers = [sheet.cell_value(0, col).strip() for col in range(sheet.ncols)]
        fn_idx = headers.index("Field Name (with ID)")
        dp_idx = headers.index("Data Point Name")
        for i in range(1, sheet.nrows):
            if str(sheet.cell_value(i, fn_idx)).strip().lower() == field_name.lower():
                return sheet.cell_value(i, dp_idx)

    elif ext == ".xlsx":
        # Use openpyxl for .xlsx
        wb = load_workbook(data_dict_path)
        sheet = wb["Data_Dictionary"]
        headers = [str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        fn_idx = headers.index("Field Name (with ID)")
        dp_idx = headers.index("Data Point Name")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[fn_idx]).strip().lower() == field_name.lower():
                return row[dp_idx]

    else:
        raise ValueError("Unsupported file type. Please provide a .xls or .xlsx file.")
    return ""

# ------------------------------------------------------------------
def get_next_ingestion_id(latest_doc_file):
    book = xlrd.open_workbook(latest_doc_file)
    sheet = book.sheet_by_index(0)
    headers = [sheet.cell_value(0, col).strip().lower().replace(" ", "") for col in range(sheet.ncols)]
    idx = next((i for i, h in enumerate(headers) if h in ['ingestionid', 'ingestionfieldid']), None)
    ade_numbers = [int(re.match(r'ADE(\d{6})', str(sheet.cell_value(i, idx))).group(1))
                   for i in range(1, sheet.nrows)
                   if re.match(r'ADE\d{6}', str(sheet.cell_value(i, idx)))]
    next_number = max(ade_numbers) + 1 if ade_numbers else 1
    return f"ADE{next_number:06d}"

def get_collaborator_association(latest_doc_file):
    book = xlrd.open_workbook(latest_doc_file)
    try:
        sheet = book.sheet_by_name("IngestionCollAssociationConfig")
    except:
        return "NA"
    headers = [sheet.cell_value(0, col).strip() for col in range(sheet.ncols)]
    idx = headers.index("Collaborator Association Reference Name")
    values = sorted(set(str(sheet.cell_value(i, idx)).strip() for i in range(1, sheet.nrows) if sheet.cell_value(i, idx)))
    values.append("NA")
    print("Collaborator Association: ")
    for i, val in enumerate(values, 1):
        print(f"{i}. {val}")
    choice = int(input("Select Collaborator Association: "))
    return values[choice - 1]

def apply_data_point_columns(abs_path, latest_doc_file, cmdd_id, ingestion_id_value,
                             user_input, source_data_point_parent,
                             collaborator_value, doc_name,
                             transformation_rule_input, default_value_input,
                             sheet_name='IngestionConfig'):

    path_parts = abs_path.strip().split('.')
    print("Select which path parts to tag with instance (optional):")
    for i, part in enumerate(path_parts):
        print(f"{i}: {part}")
    indices = input("Comma-separated indices or Enter to skip: ").strip()
    instance = input("Enter instance number (optional): ").strip()
    if indices and instance:
        idx_list = [int(i.strip()) for i in indices.split(',') if i.strip().isdigit()]
        for i in idx_list:
            path_parts[i] = f"{path_parts[i]}#Instance{instance}"

    book = xlrd.open_workbook(latest_doc_file, formatting_info=True)
    writable = xl_copy(book)
    ws = writable.get_sheet(book.sheet_names().index(sheet_name))
    sheet = book.sheet_by_name(sheet_name)
    row_idx = sheet.nrows

    row_data = ['' for _ in FIXED_HEADERS]

    def set_val(col, val):
        if col in FIXED_HEADERS:
            row_data[FIXED_HEADERS.index(col)] = val

    set_val('Ingestion id', ingestion_id_value)
    set_val('CMDD Field ID', cmdd_id)
    set_val('DataPoint-6 Source', user_input)
    set_val('Source Data Point Parent', source_data_point_parent)
    set_val('Collaborator Association', collaborator_value)
    set_val('Document Name', doc_name)
    set_val('TransformationRule', transformation_rule_input)
    set_val('Default Value', default_value_input)

    for i in range(min(6, len(path_parts) - 1)):
        set_val(f'DataPoint-{i}', path_parts[i])
    set_val('DataPoint-6', path_parts[-1])

    if path_parts[0].upper() in ['APPLICANT#INSTANCE2', 'COBORROWER']:
        set_val('DataPoint-0 DataPointContainerRule', 'ApplicantRule')

    # Fill blanks with 'NA'
    for i in range(len(row_data)):
        if not row_data[i] or str(row_data[i]).strip() == '':
            row_data[i] = 'NA'

    for col_idx, val in enumerate(row_data):
        ws.write(row_idx, col_idx, val)

    writable.save(latest_doc_file)
    print(f"✅ Data written to row {row_idx} in {latest_doc_file}")

# ─────────── Main Execution ───────────
doc_name = input("Enter Document Name (without .xls): ")
latest_doc_file = get_latest_doc_file(DOC_DIR, doc_name + ".xls")
print("Found latest file:", latest_doc_file)

user_input = input("Enter the Field Name (with ID): ")
cmdd_data = read_cmdd_xls_to_dicts(CMDD_FILE)
keyword = input("Enter keyword to search Absolute Path: ")
matches = search_absolute_path_by_keyword(keyword, cmdd_data)
cmdd_id, abs_path = select_absolute_path_from_matches(matches)

source_data_point_parent = get_source_data_point(user_input, DATA_DICT_FILE)
ingestion_id_value = get_next_ingestion_id(latest_doc_file)
collaborator_value = get_collaborator_association(latest_doc_file)
transformation_rule_input = input("Enter TransformationRule: ")
default_value_input = input("Enter Default Value: ")

apply_data_point_columns(abs_path, latest_doc_file, cmdd_id, ingestion_id_value,
                         user_input, source_data_point_parent,
                         collaborator_value, doc_name,
                         transformation_rule_input, default_value_input)
